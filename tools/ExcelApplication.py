import os
import time

import xlwings as xw
from typing import List
import pandas as pd
import numpy as np
from win32com import client
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from tools.RequestAPI import API
from dataclasses import asdict
from dacite import from_dict


class ExcelApplication:

    def __init__(self, excel_path, part_no):

        self.wb = xw.Book(excel_path)
        self.app = xw.apps.active
        self.SaveBook2XLSX(part_no)

    def ChooseSheet(self, part_no):
        return self.wb.sheets[part_no]

    def FindByKey(self, key, dataframe, return_type='value'):

        df = np.array(dataframe.values.tolist())
        for i, row in enumerate(df):
            for j, cell in enumerate(row):
                if isinstance(cell,str) and cell.strip() == key:
                    if return_type=='value':
                        return row[j+1]
                    elif return_type=='index':
                        return i, j

    def RemoveRedundant(self, dataframe, start_index: tuple):

        df = np.array(dataframe.values.tolist())
        row, col = start_index
        df = df[row:, col:]
        deleted_list = []
        deleted_col = ['S', 'L']
        for index, cell in enumerate(df[0]):
            if not cell or cell in deleted_col:
                deleted_list.append(index)
            else:
                df[0][index] = cell.strip()

        deleted_df = np.delete(df, deleted_list, 1)

        deleted_list = []
        for index, row in enumerate(deleted_df):
            if index == 0:
                continue
            if not isinstance(row[0], int):
                deleted_list.append(index)
        deleted_df = np.delete(deleted_df, deleted_list, 0)
        return deleted_df


    def SaveBook2XLSX(self, part_no):

        self.wb.save('BoomForm/{}.xlsx'.format(part_no))
        # self.wb.close()
        self.app.quit()

    def LEVEL_Normalize(self, df):

        for index, row in enumerate(df):
            if index == 0:
                continue
            text = ''
            for t in row[1]:
                if t.isdigit():
                    text += t
            df[index][1] = text
        return df

    def BoomExtreact(self, df: List[dict]):
        infos: List[dict] = []
        conserve_key = ['ITEM', 'PART NO.', 'DESCRIPTION', 'ALT.', 'GRP%' , 'QPA', 'UM', 'ITEM TEXT']

        for index, i in enumerate(df):
            if i['LEVEL'] == '1':
                if self.Is283(i['PART NO.']):
                    _283_part_end_index = self._283ExtreactIndex(df[index: ])

                    for _283_part in df[index: _283_part_end_index]:
                        d = dict()
                        for key in conserve_key:
                            d[key] = _283_part[key]
                        d['ITEM TEXT'] = '283 part'
                        infos.append(d)
                else:
                    d = dict()
                    for key in conserve_key:
                        d[key] = i[key]
                    infos.append(d)

        for i in range(len(infos)):
            infos[i]['ITEM'] = str(i+1)
        return pd.DataFrame(infos)

    def Is283(self, text):
        text = str(text)
        str_top3 = text[:3]
        if str_top3 == '283':
            return True
        else:
            return False

    def _283ExtreactIndex(self, df: List[dict]):

        for index, i in enumerate(df):
            if df[index+1]['LEVEL'] == '1':
                return index+1

    def Convert2ListOfDict(self, df):

        df = pd.DataFrame(df[1:], columns = df[0])
        d_records = np.array(df.to_dict('records'))

        return d_records

    def ExcelBoarder(self, excel_path, exclude_boarder_index):

        workbook = openpyxl.load_workbook(excel_path)

        # 取得第一個工作表
        sheet = workbook.worksheets[0]

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for i in range(1, sheet.max_row+1):

            if i < exclude_boarder_index:
                continue
            for j in range(1, sheet.max_column+1):
                sheet.cell(i, j).border = thin_border
                if i % 2 == 0:
                    sheet.cell(i, j).fill = PatternFill("solid", start_color="dbd7cc")
        workbook.save(excel_path)

    def InsertInformationCard(self, infos, excel_path):

        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.worksheets[0]

        insert_len = len([key for key in infos.keys() if not key == 'id' or not isinstance(infos[key], list)])

        sheet.insert_rows(1, insert_len)

        row_index = 1

        for key in infos.keys():
            col_index = 1

            if key == 'id' or isinstance(infos[key], list):

                if key == 'samples':
                    sample_row_index = 1

                    for sample in infos[key]:

                        sample_col_index = 4

                        if not sheet.cell(1, 4).value:

                            for sample_key in sample:
                                sheet.cell(sample_row_index, sample_col_index).value = sample_key
                                sample_col_index += 1
                            sample_row_index += 1

                        sample_col_index = 4
                        for value in sample.values():

                            sheet.cell(sample_row_index, sample_col_index).value = value
                            sample_col_index += 1
                        sample_row_index += 1

                elif key == 'remarks':

                    sheet.insert_rows(insert_len, len(infos[key]))
                    insert_len += len(infos[key])
                    sheet.cell(7, 3).value = '備註'
                    index = 1
                    for value in infos[key]:

                        sheet.cell(7+index, 3).value = value
                        index += 1

                else:
                    continue

            else:
                sheet.cell(row_index, col_index).value = key
                sheet.cell(row_index, col_index+1).value = infos[key]
                row_index += 1

        workbook.save(excel_path)

        return insert_len

    def Convert2PDF(self, excel_path, output_path):

        excel = client.Dispatch('Excel.Application')
        wb = excel.Workbooks.Open(excel_path)
        ws = wb.Worksheets(1)
        ws.Columns.AutoFit()
        ws.PageSetup.Zoom = False
        ws.PageSetup.FitToPagesTall = 1
        ws.PageSetup.FitToPagesWide = 1
        ws.PageSetup.LeftMargin = 0
        ws.PageSetup.RightMargin = 0
        ws.PageSetup.TopMargin = 0
        ws.PageSetup.BottomMargin = 0

        wb.SaveAs(excel_path)

        pdf_path = excel_path.replace('xlsx', 'pdf')

        wb.ActiveSheet.ExportAsFixedFormat(0, pdf_path)
        wb.Close()

        excel.Application.Quit()

        return pdf_path

def ExcelOperator(path, part_no, data):

    excel_application = ExcelApplication(path, part_no)
    df = pd.read_excel('BoomForm/{}.xlsx'.format(part_no))

    df = df.fillna('')
    date = excel_application.FindByKey('DATE:', df)
    rev = str(int(excel_application.FindByKey('REV:', df))).zfill(2)
    item_index = excel_application.FindByKey('ITEM', df, 'index')
    boom_ = excel_application.RemoveRedundant(df, item_index)
    boom_ = excel_application.LEVEL_Normalize(boom_)
    boom_ = excel_application.Convert2ListOfDict(boom_)
    infos = excel_application.BoomExtreact(boom_)
    save_excel_path = os.path.join(os.getcwd(), 'BoomForm/{}.xlsx'.format(part_no))
    infos.to_excel(save_excel_path, index = False)

    infos = data
    infos.date = date
    infos.rev = rev

    infos_dict = asdict(infos)
    exclude_boarder_index = excel_application.InsertInformationCard(infos_dict, save_excel_path)
    excel_application.ExcelBoarder(save_excel_path, exclude_boarder_index)
    pdf_path = excel_application.Convert2PDF(save_excel_path, '')

    return pdf_path, infos

if __name__ == '__main__':
    part_no = '2870676303'
    path = r'D:\DeltaBox\OneDrive - Delta Electronics, Inc\文件\SAP\SAP GUI\{}.xls'.format(part_no)
    ExcelOperator(path, part_no)
